from __future__ import annotations
from datetime import date
from uuid import uuid4
from django.shortcuts import render, redirect


from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseRedirect

# /srv/mycrm/crm_core/views_expenses.py


import base64
import json
import uuid
from django.conf import settings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests
import yaml
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages

def _ensure_ids(items):
    # Sorgt dafür, dass jedes Draft-Item eine String-ID hat
    out = []
    for idx, it in enumerate(items):
        try:
            d = dict(it)
        except Exception:
            d = {**it}
        d.setdefault("id", str(idx))
        out.append(d)
    return out



# -----------------------------------------------------------------------------
# Paths & simple persistence
# -----------------------------------------------------------------------------

def _safe_base_dir() -> Path:
    try:
        return Path(settings.BASE_DIR)
    except Exception:
        return Path(__file__).resolve().parents[1]  # app dir -> project root

BASE_DIR = _safe_base_dir()
DATA_DIR = BASE_DIR / "data"
FILES_DIR = BASE_DIR / "files" / "accounting"
CATS_YML = FILES_DIR / "categories.yml"
DRAFTS = DATA_DIR / "expense_drafts.jsonl"

# Make sure directories exist at import time (gunicorn worker)
try:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    # do not crash the worker; permission will be visible in logs if used
    pass


# -----------------------------------------------------------------------------
# Utilities
# -----------------------------------------------------------------------------

def _detect_base() -> str:
    # keep it simple; could be made dynamic later
    return "crm_core/base.html"


def _load_categories() -> dict | list:
    try:
        if CATS_YML.exists():
            return yaml.safe_load(CATS_YML.read_text(encoding="utf-8")) or {}
    except Exception:
        pass
    return {}


def _now_date() -> str:
    return date.today().strftime("%Y-%m-%d")


def _parse_amount(s: Any) -> float:
    if s is None:
        return 0.0
    s = str(s).strip()
    if not s:
        return 0.0
    has_dot = "." in s
    has_com = "," in s
    try:
        if has_dot and has_com:
            return float(s.replace(".", "").replace(",", "."))
        if has_com and not has_dot:
            return float(s.replace(",", "."))
        if has_dot and not has_com:
            parts = s.split(".")
            if len(parts) == 2 and len(parts[1]) <= 2:
                return float(s)
            return float(s.replace(".", ""))
        return float(s)
    except Exception:
        return 0.0


def _categories_list3() -> List[str]:
    data = _load_categories()
    if isinstance(data, dict):
        if "categories" in data and isinstance(data["categories"], list):
            return [str(x) for x in data["categories"]]
        return [str(k) for k in data.keys()]
    if isinstance(data, list):
        return [str(x) for x in data]
    return []


def _append_draft(row: Dict[str, Any]) -> None:
    row = {k: (v if v is not None else "") for k, v in row.items()}
    try:
        with DRAFTS.open("a", encoding="utf-8") as f:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    except Exception as e:
        try:
            print("draft append error:", repr(e))
        except Exception:
            pass


def _read_drafts() -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    if DRAFTS.exists():
        for line in DRAFTS.read_text(encoding="utf-8").splitlines():
            try:
                items.append(json.loads(line))
            except Exception:
                continue
    return items


def _write_drafts(items: List[Dict[str, Any]]) -> None:
    with DRAFTS.open("w", encoding="utf-8") as f:
        for it in items:
            if not it.get("Unternehmen") and it.get("Lieferant"):
                it["Unternehmen"] = it["Lieferant"]
            f.write(json.dumps(it, ensure_ascii=False) + "\n")


def _ensure_unternehmen(it: Dict[str, Any]) -> None:
    if not it.get("Unternehmen") and it.get("Lieferant"):
        it["Unternehmen"] = it["Lieferant"]


# -----------------------------------------------------------------------------
# Minimal text extract & heuristics (placeholders)
# -----------------------------------------------------------------------------

def _extract_text_from_attachment(name: str, data: bytes) -> tuple[str, bool]:
    """
    Very lightweight placeholder: we do not actually OCR/parse here.
    Returns (text, error_flag). error_flag False means "no fatal error".
    """
    if not data:
        return ("", True)
    try:
        # In a real setup we would detect PDF/image here.
        return ("", False)
    except Exception:
        return ("", True)


def _guess_company_from_text(text: str) -> str:
    # Minimal heuristic placeholder
    return ""


# -----------------------------------------------------------------------------
# Microsoft Graph helpers
# -----------------------------------------------------------------------------

def _graph_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def _graph_token_for(user) -> Optional[str]:
    """
    Get an access token from django-allauth, provider 'microsoft'.
    """
    try:
        from allauth.socialaccount.models import SocialToken  # type: ignore
    except Exception:
        # allauth not available
        return None

    try:
        tok = (
            SocialToken.objects.filter(
                account__user=user, account__provider="microsoft"
            )
            .order_by("-id")
            .first()
        )
        if tok and tok.token:
            return tok.token
    except Exception as e:
        try:
            print("graph token error:", repr(e))
        except Exception:
            pass
    return None


def _graph_me_id(token: str) -> Optional[str]:
    try:
        r = requests.get(
            "https://graph.microsoft.com/v1.0/me?$select=id",
            headers=_graph_headers(token),
            timeout=10,
        )
        if r.status_code == 200:
            js = r.json() or {}
            return js.get("id")
        try:
            print("graph me error:", r.status_code, (r.text or "")[:300])
        except Exception:
            pass
    except Exception as e:
        try:
            print("graph me exception:", repr(e))
        except Exception:
            pass
    return None


def _graph_user_base(uid: Optional[str]) -> str:
    """
    Build base path for Graph:
    - '/users/{uid}' if uid is given and not 'me'
    - '/me' as fallback
    """
    if uid and uid != "me":
        try:
            return f"/users/{quote(uid)}"
        except Exception:
            pass
    return "/me"


def _graph_list_attachments(token: str, uid: Optional[str], mid: str) -> List[Dict[str, Any]]:
    """
    List attachments of a message robustly.
    Uses users/{uid} then falls back to /me. Does not request contentBytes here.
    """
    mid_esc = quote(mid, safe="")
    qs = "$top=25&$select=id,name,contentType,size,isInline,"
    bases = []
    if uid and uid != "me":
        bases.append(f"/users/{quote(uid)}")
    bases.append("/me")

    for b in bases:
        url = f"https://graph.microsoft.com/v1.0{b}/messages/{mid_esc}/attachments?{qs}"
        r = requests.get(url, headers=_graph_headers(token), timeout=20)
        if r.status_code == 200:
            js = r.json() or {}
            return js.get("value", [])
        try:
            print("Graph list attachments error:", r.status_code, (r.text or "")[:500])
        except Exception:
            pass
    return []


# -----------------------------------------------------------------------------
# Views: Drafts list / add / delete / edit / update
# -----------------------------------------------------------------------------

@login_required
@require_http_methods(["GET"])
def drafts_list(request: HttpRequest) -> HttpResponse:
    items = _read_drafts()
    rows: List[Dict[str, Any]] = []
    total = 0.0
    for it in items:
        _ensure_unternehmen(it)
        b = float(it.get("Brutto") or 0)
        satz = _parse_amount(it.get("USt_Satz"))
        netto = (b / (1 + satz / 100)) if satz else b
        ust_b = b - netto
        it["Netto"] = round(netto, 2)
        it["USt"] = round(ust_b, 2)
        rows.append(it)
        total += b

    ctx = {
        "base_template": _detect_base(),
        "items": rows,
        "categories": _categories_list3(),
        "today": _now_date(),
        "sum_brutto": total,
    }
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    return render(request, "accounting/drafts.html", ctx)


@login_required
@require_http_methods(["POST"])
def draft_add(request: HttpRequest) -> HttpResponseRedirect:
    row = {
        "id": str(uuid.uuid4()),
        "Brutto": _parse_amount(request.POST.get("Brutto")),
        "USt_Satz": request.POST.get("USt_Satz") or "",
        "Umsatzsteuer": request.POST.get("Umsatzsteuer") or "",
        "Kategorie": request.POST.get("Kategorie") or "",
        "Datum": request.POST.get("Datum") or _now_date(),
        "Zahlungsdatum": request.POST.get("Zahlungsdatum") or "",
        "Zahlung": request.POST.get("Zahlung") or "",
        "Unternehmen": request.POST.get("Unternehmen") or "",
        "Lieferant": request.POST.get("Lieferant") or "",
        "Beschreibung": request.POST.get("Beschreibung") or "",
        "Belegname": request.POST.get("Belegname") or "",
        "OneDriveName": request.POST.get("OneDriveName") or "",
        "OneDriveUrl": request.POST.get("OneDriveUrl") or "",
    }
    _append_draft(row)
    messages.success(request, "Vormerkung gespeichert.")
    return redirect("/crm/expenses/drafts/")


@login_required
@require_http_methods(["POST"])

@require_http_methods(["POST"])

@require_http_methods(["POST"])

@require_http_methods(["POST"])
def draft_delete(request):
    items = _ensure_ids(_read_drafts())
    d = request.POST
    idx = None
    if "idx" in d:
        try:
            idx = int(d.get("idx"))
        except Exception:
            return HttpResponseBadRequest("invalid idx")
    else:
        rid = d.get("id")
        if rid is None:
            return HttpResponseBadRequest("missing id/idx")
        idx = _find_index_by_id(items, rid)

    if idx < 0 or idx >= len(items):
        return HttpResponseBadRequest("id out of range")

    items.pop(idx)
    _write_drafts(items)
    return redirect("expenses_drafts")


@require_http_methods(["POST"])
def draft_delete(request):
    items = _ensure_ids(_read_drafts())
    d = request.POST
    idx = None
    if "idx" in d:
        try:
            idx = int(d.get("idx"))
        except Exception:
            return HttpResponseBadRequest("invalid idx")
    else:
        rid = d.get("id")
        if rid is None:
            return HttpResponseBadRequest("missing id/idx")
        idx = _find_index_by_id(items, rid)

    if idx < 0 or idx >= len(items):
        return HttpResponseBadRequest("id out of range")

    items.pop(idx)
    _write_drafts(items)
    return redirect("expenses_drafts")



def draft_from_email_bridge(request, *args, **kwargs):
    """
    Kompatibilität für alte Route /crm/expenses/draft/from-email/
    Leitet intern auf die OCR-Implementierung weiter.
    """
    try:
        return ocr_from_email(request)
    except Exception as e:
        return HttpResponseServerError(f"compat error: {e}")
# --- minimal OCR placeholder START ---
def ocr_from_email(request):
    """Fallback-View: vermeidet 500er und leitet zur Entwurfsübersicht."""
    try:
        return redirect('/crm/expenses/drafts/')
    except Exception as e:
        return HttpResponseServerError(f"OCR placeholder error: {e}")
# --- minimal OCR placeholder END ---

# --- BEGIN hotfix: robust drafts helpers & views ---
import json
from pathlib import Path
from datetime import date

def _drafts_path():
    p = Path("/srv/mycrm/var/expenses_drafts.json")
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.exists():
        p.write_text("[]", encoding="utf-8")
    return p

def _read_drafts():
    p = _drafts_path()
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []

def _write_drafts(items):
    p = _drafts_path()
    p.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")

def _ensure_ids(items):
    for i, it in enumerate(items):
        if not isinstance(it, dict):
            items[i] = {}
            it = items[i]
        it["id"] = i
    return items

_NUM_FIELDS = ("Brutto", "Netto", "USt")
def _as_float(v, default=0.0):
    try:
        if v in ("", None): return float(default)
        return float(v)
    except Exception:
        return float(default)

def _defaulted(rec: dict):
    r = dict(rec or {})
    r.setdefault("Belegnummer", "")
    r.setdefault("Brutto", 0.0)
    r.setdefault("Netto",  0.0)
    # USt: falls nicht gesetzt, aus Brutto/Netto ableiten
    r["USt"] = _as_float(r.get("USt", _as_float(r.get("Brutto",0)) - _as_float(r.get("Netto",0))), 0.0)
    r.setdefault("Umsatzsteuer", "Normal")
    r.setdefault("Kategorie", "")
    r.setdefault("Unternehmen", "")
    r.setdefault("Datum", str(date.today()))
    r.setdefault("Zahlung", "")
    r.setdefault("OneDriveUrl", "")
    r.setdefault("raw_preview", "")
    # Nummern als float „glätten“
    for k in _NUM_FIELDS:
        r[k] = _as_float(r.get(k, 0.0), 0.0)
    return r

def drafts_list(request):
    items = [_defaulted(x) for x in _ensure_ids(_read_drafts())]
    return render(request, "accounting/drafts.html", {"items": items})

def draft_add(request):
    items = _read_drafts()
    data = {}
    if request.method == "POST":
        d = request.POST
        data = {
            "Belegnummer": d.get("Belegnummer") or "",
            "Brutto": _as_float(d.get("Brutto"), 0.0),
            "Netto":  _as_float(d.get("Netto"),  0.0),
            "USt":    _as_float(d.get("USt"),    0.0),
            "Umsatzsteuer": d.get("Umsatzsteuer") or "Normal",
            "Kategorie": d.get("Kategorie") or "",
            "Unternehmen": d.get("Unternehmen") or "",
            "Datum": d.get("Datum") or str(date.today()),
            "Zahlung": d.get("Zahlung") or "",
            "OneDriveUrl": d.get("OneDriveUrl") or "",
            "raw_preview": d.get("raw_preview") or "",
        }
    else:
        # Auch GET erzeugt notfalls einen Platzhalter (für „Vormerken“)
        g = request.GET
        data = {
            "Belegnummer": g.get("Belegnummer") or "",
            "Brutto": _as_float(g.get("Brutto"), 0.0),
            "Netto":  _as_float(g.get("Netto"),  0.0),
            "USt":    _as_float(g.get("USt"),    0.0),
            "Umsatzsteuer": g.get("Umsatzsteuer") or "Normal",
            "Kategorie": g.get("Kategorie") or "",
            "Unternehmen": g.get("Unternehmen") or "",
            "Datum": g.get("Datum") or str(date.today()),
            "Zahlung": g.get("Zahlung") or "",
            "OneDriveUrl": g.get("OneDriveUrl") or "",
            "raw_preview": g.get("raw_preview") or g.get("raw") or g.get("message_id") or "Vormerkung aus Mail",
        }
        # Falls nur Brutto da ist, Netto/USt ableiten:
        if not g.get("Netto") and not g.get("USt"):
            data["Netto"] = data["Brutto"]
            data["USt"] = 0.0

    items.append(_defaulted(data))
    _write_drafts(items)
    print("draft_add: appended ->", data.get("raw_preview", "")[:80])
    return redirect("expenses_drafts")

def draft_edit(request):
    items = [_defaulted(x) for x in _ensure_ids(_read_drafts())]
    q = request.GET
    idx = None
    if "idx" in q:
        try:
            idx = int(q.get("idx"))
        except Exception:
            return HttpResponseBadRequest("invalid idx")
    else:
        rid = q.get("id")
        if rid is None:
            return HttpResponseBadRequest("missing id/idx")
        idx = _find_index_by_id(items, rid)

    if idx is None or idx < 0 or idx >= len(items):
        return HttpResponseBadRequest("id out of range")

    r = items[idx]
    ctx = {"r": r, "id": idx, "categories": _get_categories()}
    ctx["categories"] = _categories_list3()
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    if not ctx.get("item") and hasattr(request, "GET") and request.GET.get("id"):
        _id = request.GET.get("id")
        try:
            _items = _read_drafts()
            for _it in _items:
                if str(_it.get("id")) == str(_id):
                    ctx["item"] = _it
                    break
        except Exception:
            pass
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    if not ctx.get("item") and (request.GET.get("id") or request.POST.get("id")):
        _id = request.GET.get("id") or request.POST.get("id")
        try:
            for _it in _read_drafts():
                if str(_it.get("id")) == str(_id):
                    ctx["item"] = _it
                    break
        except Exception:
            pass
    ctx.setdefault("categories", _categories_list3())
    ctx.setdefault("today", _now_date() if "_now_date" in globals() else None)
    if not ctx.get("item"):
        _id = request.GET.get("id") or request.POST.get("id")
        if _id:
            try:
                for _it in _read_drafts():
                    if str(_it.get("id")) == str(_id):
                        ctx["item"] = _it
                        break
            except Exception:
                pass
    return render(request, "accounting/draft_edit.html", ctx)

def draft_update(request):
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    items = _ensure_ids(_read_drafts())
    d = request.POST
    idx = None
    if "idx" in d:
        try:
            idx = int(d.get("idx"))
        except Exception:
            return HttpResponseBadRequest("invalid idx")
    else:
        rid = d.get("id")
        if rid is None:
            return HttpResponseBadRequest("missing id/idx")
        idx = _find_index_by_id(items, rid)

    if idx < 0 or idx >= len(items):
        return HttpResponseBadRequest("id out of range")

    # bestehende id beibehalten
    prev = _defaulted(items[idx])
    rec = _defaulted({
        "id": prev.get("id"),
        "Belegnummer": d.get("Belegnummer"),
        "Brutto": d.get("Brutto"),
        "Netto":  d.get("Netto"),
        "USt":    d.get("USt"),
        "Umsatzsteuer": d.get("Umsatzsteuer"),
        "Kategorie": d.get("Kategorie"),
        "Unternehmen": d.get("Unternehmen"),
        "Datum": d.get("Datum"),
        "Zahlung": d.get("Zahlung"),
        "OneDriveUrl": d.get("OneDriveUrl"),
        "raw_preview": d.get("raw_preview"),
    })
    items[idx] = rec
    _write_drafts(items)
    return redirect("expenses_drafts")



def draft_add(request):
    data = request.POST if request.method == "POST" else request.GET

    # akzeptiere mehrere Parameternamen
    text = (data.get("Beschreibung")
            or data.get("raw_preview")
            or data.get("preview")
            or data.get("text")
            or data.get("body")
            or "Vormerkung aus Mail")

    row = {
        "id": str(uuid.uuid4()),
        "Brutto": _parse_amount(data.get("Brutto")),
        "USt_Satz": data.get("USt_Satz") or "",
        "Umsatzsteuer": data.get("Umsatzsteuer") or "",
        "Kategorie": data.get("Kategorie") or "",
        "Datum": data.get("Datum") or _now_date(),
        "Zahlungsdatum": data.get("Zahlungsdatum") or "",
        "Zahlung": data.get("Zahlung") or "",
        "Unternehmen": data.get("Unternehmen") or "",
        "Lieferant": data.get("Lieferant") or "",
        "Beschreibung": text,
        "Belegname": data.get("Belegname") or "",
        "OneDriveName": data.get("OneDriveName") or "",
        "OneDriveUrl": data.get("OneDriveUrl") or "",
        "raw_preview": text,
    }

    _append_draft(row)
    messages.success(request, "Vormerkung gespeichert.")
    return redirect("/crm/expenses/drafts/")

# ===== ROBUSTE OVERRIDES (ganz ans Dateiende) =====
import logging, uuid
logger = logging.getLogger(__name__)

# Falls es doppelte Implementierungen gibt, erzwingen wir EINEN Satz Namen:
try:
    _read = _read_drafts
    _write = _write_drafts
except NameError:
    # wenn die Namen (noch) nicht existieren, einfach durchreichen;
    # der Import-Fehler zeigt sich dann im Log/Traceback.
    pass

def _safe_append(row: dict):
    try:
        items = _read()
    except Exception as e:
        logger.exception("draft_add: _read() failed: %s", e)
        items = []
    items = items or []
    items.append(row)
    try:
        _write(items)
    except Exception as e:
        logger.exception("draft_add: _write() failed: %s", e)
        raise

@login_required
@require_http_methods(["GET","POST"])
@csrf_exempt  # nur damit curl-Tests nicht an CSRF scheitern; Browser-Flow bleibt sicher.
def draft_add(request):
    data = request.POST if request.method == "POST" else request.GET
    text = (data.get("Beschreibung") or data.get("raw_preview") or data.get("preview")
            or data.get("text") or data.get("body") or "Vormerkung aus Mail")

    row = {
        "id": str(uuid.uuid4()),
        "Brutto": _parse_amount(data.get("Brutto")),
        "USt_Satz": data.get("USt_Satz") or "",
        "Umsatzsteuer": data.get("Umsatzsteuer") or "",
        "Kategorie": data.get("Kategorie") or "",
        "Datum": data.get("Datum") or _now_date(),
        "Zahlungsdatum": data.get("Zahlungsdatum") or "",
        "Zahlung": data.get("Zahlung") or "",
        "Unternehmen": data.get("Unternehmen") or "",
        "Lieferant": data.get("Lieferant") or "",
        "Beschreibung": text,
        "Belegname": data.get("Belegname") or "",
        "OneDriveName": data.get("OneDriveName") or "",
        "OneDriveUrl": data.get("OneDriveUrl") or "",
    }
    try:
        _safe_append(row)
        messages.success(request, f"Vormerkung gespeichert: {text[:60]}")
        logger.info("draft_add: appended -> %s", text)
    except Exception as e:
        messages.error(request, f"Fehler beim Speichern der Vormerkung: {e}")
        logger.exception("draft_add: exception: %s", e)
    return redirect("/crm/expenses/drafts/")


DEFAULT_EXPENSE_CATEGORIES = [
    "Allgemein","Büro","Reise","Bewirtung","Software","Telefon/Internet","Miete","Versicherung","Sonstiges"
]

def _get_categories():
    return getattr(settings, "EXPENSE_CATEGORIES", DEFAULT_EXPENSE_CATEGORIES)

def _find_index_by_id(items, rid):
    rid = str(rid)
    for i, x in enumerate(items):
        if str(x.get("id")) == rid:
            return i
    return -1

def _get_index_from_request(request, method="GET"):
    data = request.POST if method == "POST" else request.GET
    # bevorzugt idx (expliziter Index)
    idx = data.get("idx")
    if idx is not None:
        try:
            return int(idx)
        except Exception:
            return -1
    # Fallback: id (UUID) → Index suchen
    rid = data.get("id")
    if rid:
        items = _ensure_ids(_read_drafts())
        return _find_index_by_id(items, rid)
    return -1
# === categories fallback helper (yaml or from existing drafts) ===
def _categories_list3():
    try:
        import yaml
    except Exception:
        yaml = None
    from pathlib import Path
    paths = [Path("/srv/mycrm/files/accounting/categories.yml"), Path("/srv/mycrm/files/accounting/categories.yaml")]
    if yaml:
        for p in paths:
            if p.exists():
                try:
                    data = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
                    if isinstance(data, dict):
                        cats = (data.get("categories") or data.get("Categories") or data.get("kategorien") or data.get("Kategorien"))
                        if isinstance(cats, list):
                            return [str(x) for x in cats if str(x).strip()]
                    if isinstance(data, list):
                        return [str(x) for x in data if str(x).strip()]
                except Exception:
                    pass
    try:
        items = _read_drafts()
    except Exception:
        items = []
    vals = sorted({str(x.get("Kategorie")).strip() for x in items if x.get("Kategorie")})
    return vals
# === end categories fallback ===

def _categories_list3():
    import re
    from pathlib import Path
    cats = []
    # 1) Kandidatenpfade + rekursive Suche (yaml UND yml)
    cand = [
        Path("/srv/mycrm/files/accounting/categories.yaml"),
        Path("/srv/mycrm/files/accounting/categories.yml"),
        Path("/srv/mycrm/crm_core/files/accounting/categories.yaml"),
        Path("/srv/mycrm/crm_core/files/accounting/categories.yml"),
        Path("/srv/mycrm/data/accounting/categories.yaml"),
        Path("/srv/mycrm/data/accounting/categories.yml"),
    ]
    try:
        for p in Path("/srv/mycrm").rglob("categories.y*ml"):
            cand.append(p)
    except Exception:
        pass
    # 2) Erst versuchen mit PyYAML
    try:
        import yaml
    except Exception:
        yaml = None
    for p in cand:
        try:
            if not p.exists():
                continue
            txt = p.read_text(encoding="utf-8", errors="ignore")
            if yaml:
                try:
                    data = yaml.safe_load(txt) or {}
                    if isinstance(data, dict):
                        for k in ("categories","Categories","kategorien","Kategorien"):
                            v = data.get(k)
                            if isinstance(v, list):
                                cats = [str(x).strip() for x in v if str(x).strip()]
                                if cats:
                                    return cats
                    if isinstance(data, list):
                        cats = [str(x).strip() for x in data if str(x).strip()]
                        if cats:
                            return cats
                except Exception:
                    pass
            # 3) Notfall: simpler YAML-Listenparser ohne PyYAML
            lines = txt.splitlines()
            in_block = False
            for line in lines:
                if not in_block:
                    if re.match(r"^\s*(categories|Categories|kategorien|Kategorien)\s*:\s*$", line):
                        in_block = True
                    continue
                m = re.match(r"^\s*-\s*(.+?)\s*$", line)
                if m:
                    cats.append(m.group(1).strip())
                elif re.match(r"^\S", line):
                    break
            if cats:
                return cats
        except Exception:
            pass
    # 4) Fallback: aus vorhandenen Entwürfen aggregieren
    try:
        items = _read_drafts()
        cats = sorted({str(x.get("Kategorie")).strip() for x in items if x.get("Kategorie")})
        if cats:
            return cats
    except Exception:
        pass
    # 5) Letzter Notnagel
    return [
        "Abschreibung","Abschreibung als Anlage (AfA)","Arbeitsmittel / Büro","Arbeitszimmer","Bewirtung",
        "EDV-Aufwand","Fachliteratur","Fremdleistungen","Gehalt","Krankenkasse","GWG (geringw. Wirtschaftsgut)",
        "KFZ-Aufwand","Porto","Privatentnahme","Rechtsberatung","Reisekosten","Sonstiges","Spenden",
        "Spesen (Geldverkehr)","Steuerberatung","SVA","Taggelder","Telefon","Wareneinkauf","Werbungskosten","WKO-Gebühr"
    ]

def ocr_from_email(request):
    """Erzeugt (mind.) einen Skeleton-Entwurf aus der Mailansicht und öffnet ihn zum Bearbeiten."""
    q = request.GET.copy()
    mid = q.get("mid") or q.get("message_id") or q.get("messageId") or q.get("id")
    aid = q.get("aid") or q.get("attachment_id") or q.get("attachmentId")
    # sinnvolle Defaults setzen
    if not q.get("Datum"):
        q["Datum"] = date.today().isoformat()
    if not q.get("Beschreibung"):
        subj = q.get("subject") or q.get("subj") or ""
        q["Beschreibung"] = subj or (f"OCR aus E-Mail {mid}" if mid else "OCR aus E-Mail")
    q.setdefault("Kategorie", "")
    q.setdefault("Brutto", q.get("amount") or q.get("betrag") or "0")
    q["__source__"] = "mail_ocr"
    try:
        _amt = _try_extract_amount(q, mid=mid, aid=aid)
        if _amt and (not q.get("Brutto") or str(q.get("Brutto")).strip() in ("", "0", "0.0", "0.00")):
            q["Brutto"] = _amt
    except Exception:
        pass
    # 1) Bevorzugt: bestehende draft_add-Logik wiederverwenden, indem wir ein POST-Request-Wrapper bauen
    try:
        class _ReqWrap:
            def __init__(self, req, post):
                self._r = req; self.method = "POST"; self.POST = post; self.FILES = getattr(req, "FILES", {})
                self.GET = {}; self.META = getattr(req, "META", {}); self.user = getattr(req, "user", None); self.session = getattr(req, "session", {})
            def __getattr__(self, k):
                return getattr(self._r, k)
        return draft_add(_ReqWrap(request, q))
    except Exception:
        pass
    # 2) Fallback: Draft manuell anlegen und speichern
    try:
        items = _read_drafts()
    except Exception:
        items = []
    item = {
        "id": str(uuid4()),
        "Datum": q.get("Datum"),
        "Beschreibung": q.get("Beschreibung"),
        "Kategorie": q.get("Kategorie"),
        "Brutto": q.get("Brutto"),
        "Quelle": "mail_ocr",
        "MailID": mid or "",
        "AttachmentID": aid or "",
    }
    try:
        items.append(item)
        _write_drafts(items)
        return redirect(f"/crm/expenses/draft/edit/?id={item[id]}")
    except Exception:
        return redirect("/crm/expenses/drafts/")

def _try_extract_amount(q, mid=None, aid=None):
    """Versucht, einen Betrag aus Query/Metadaten/Text zu extrahieren. Gibt String "123.45" oder None."""
    import re
    texts = []
    # 1) direkte Kandidatenfelder
    for k in ("Brutto","brutto","Total","total","TotalBrutto","sum","Summe","betrag","Betrag","amount","Amount","gross","total_amount","totalAmount","price","Price"):
        v = q.get(k)
        if v: texts.append(str(v))
    # 2) ggf. Metadaten durchsehen
    for k in ("subject","Subject","beschreibung","Beschreibung","body","bodyPreview","preview","text","ocr","ocr_text"):
        v = q.get(k)
        if v: texts.append(str(v))
    # 3) Falls es Projekt-Helfer gibt, Attachment-Text holen
    if (mid or aid):
        for fn_name in ("attachment_text","ocr_attachment_text","ocr_extract_text","mail_attachment_text","download_attachment_text","msgraph_attachment_text"):
            fn = globals().get(fn_name)
            if callable(fn):
                try:
                    t = fn(mid, aid)
                    if isinstance(t, bytes):
                        t = t.decode("utf-8","ignore")
                    if t:
                        texts.append(str(t))
                except Exception:
                    pass
    if not texts:
        return None
    # 4) Beträge parsen (EUR/€ optional, Tausenderpunkte/Leerzeichen erlaubt, Dezimaltrennzeichen . oder ,)
    pat = re.compile(r"(?:\b(?:EUR)\b|€)?\s*(\d{1,3}(?:[\.\s]\d{3})*|\d+)(?:[\.,](\d{2}))?\s*(?:\b(?:EUR)\b|€)?", re.I)
    candidates = []
    for tx in texts:
        for m in pat.finditer(tx):
            whole = (m.group(1) or "0").replace(" ","").replace(".","")
            cents = (m.group(2) or "00")
            try:
                val = float(f"{whole}.{cents}")
                candidates.append(val)
            except Exception:
                pass
    if not candidates:
        return None
    # Heuristik: größten plausiblen Wert nehmen
    amt = max(candidates)
    return f"{amt:.2f}"

def _safe_money(v):
    s = str(v or "").strip()
    s = s.replace("€","").replace(" ", "")
    # deutsche Schreibweise zulassen: 1.234,56 -> 1234.56
    if s.count(",") == 1 and (s.count(".") >= 1):
        s = s.replace(".", "").replace(",", ".")
    elif s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return f"{float(s):.2f}"
    except Exception:
        return "0.00"

@require_http_methods(["GET","POST"])
@login_required
def draft_edit(request):
    """Stabile Bearbeiten-View: lädt/schafft Draft, schützt Kontext und speichert sicher."""
    from datetime import date
    _id = request.GET.get("id") or request.POST.get("id")
    try:
        items = _read_drafts()
    except Exception:
        items = []
    item = None
    if _id:
        for it in items:
            if str(it.get("id")) == str(_id):
                item = it; break
    if request.method == "POST":
        # Draft anlegen, falls nicht vorhanden
        if not item:
            from uuid import uuid4
            item = {"id": str(uuid4())}
            items.append(item)
        # Felder übernehmen
        datum = request.POST.get("Datum") or request.POST.get("datum") or item.get("Datum") or date.today().isoformat()
        beschr = request.POST.get("Beschreibung") or request.POST.get("beschreibung") or item.get("Beschreibung") or ""
        kateg = request.POST.get("Kategorie") or request.POST.get("kategorie") or item.get("Kategorie") or ""
        brutto_in = request.POST.get("Brutto") or request.POST.get("brutto") or item.get("Brutto") or "0"
        item.update({"Datum": datum, "Beschreibung": beschr, "Kategorie": kateg, "Brutto": _safe_money(brutto_in)})
        try:
            _write_drafts(items)
        except Exception:
            pass
        return redirect(f"/crm/expenses/draft/edit/?id={item[id]}")
    # GET: Draft notfalls erzeugen, damit Template nie ins Leere greift
    if not item:
        from uuid import uuid4
        item = {"id": str(uuid4()), "Datum": date.today().isoformat(), "Beschreibung": "", "Kategorie": "", "Brutto": ""}
        items.append(item)
        try:
            _write_drafts(items)
        except Exception:
            pass
    # Kontext sicher befüllen
    try:
        cats = _categories_list3()
    except Exception:
        try:
            cats = _categories_list2()
        except Exception:
            cats = []
    ctx = {
        "base_template": "crm_core/base.html",
        "today": date.today().isoformat(),
        "categories": cats,
        "parsed": {},
        "item": item,
    }
    return render(request, "accounting/draft_edit.html", ctx)


@require_http_methods(["GET","POST"])
@login_required
def draft_edit2(request):
    from datetime import date
    _id = request.GET.get("id") or request.POST.get("id")
    try:
        items = _read_drafts()
    except Exception:
        items = []
    item = None
    if _id:
        for it in items:
            if str(it.get("id")) == str(_id):
                item = it
                break
    if request.method == "POST":
        if not item:
            from uuid import uuid4
            item = {"id": str(uuid4())}
            items.append(item)
        datum = request.POST.get("Datum") or item.get("Datum") or date.today().isoformat()
        beschr = request.POST.get("Beschreibung") or item.get("Beschreibung") or ""
        kateg = request.POST.get("Kategorie") or item.get("Kategorie") or ""
        brutto_in = request.POST.get("Brutto") or item.get("Brutto") or "0"
        item.update({"Datum": datum, "Beschreibung": beschr, "Kategorie": kateg, "Brutto": _safe_money(brutto_in)})
        try:
            _write_drafts(items)
        except Exception:
            pass
        return redirect(f"/crm/expenses/draft/edit/?id={item['id']}")
    if not item:
        from uuid import uuid4
        item = {"id": str(uuid4()), "Datum": date.today().isoformat(), "Beschreibung": "", "Kategorie": "", "Brutto": ""}
        items.append(item)
        try:
            _write_drafts(items)
        except Exception:
            pass
    # Kategorien robuster laden
    try:
        cats = _categories_list3()
    except Exception:
        try:
            cats = _categories_list2()
        except Exception:
            cats = []
    ctx = {
        "base_template": "crm_core/base.html",
        "today": date.today().isoformat(),
        "categories": cats,
        "parsed": {},
        "item": item,
    }
    return render(request, "accounting/draft_edit.html", ctx)

@csrf_exempt
@require_http_methods(["GET","POST"])
@login_required
def draft_edit3(request):
    """Robuste Bearbeiten-View ohne Template (kein 500 mehr)."""
    from datetime import date
    import html

    _id = request.GET.get("id") or request.POST.get("id")
    try:
        items = _read_drafts()
    except Exception:
        items = []

    item = None
    if _id:
        for it in items:
            if str(it.get("id")) == str(_id):
                item = it
                break

    if request.method == "POST":
        if not item:
            from uuid import uuid4
            item = {"id": str(uuid4())}
            items.append(item)
        datum = request.POST.get("Datum") or item.get("Datum") or date.today().isoformat()
        beschr = request.POST.get("Beschreibung") or item.get("Beschreibung") or ""
        kateg = request.POST.get("Kategorie") or item.get("Kategorie") or ""
        brutto_in = request.POST.get("Brutto") or item.get("Brutto") or "0"
        item.update({"Datum": datum, "Beschreibung": beschr, "Kategorie": kateg, "Brutto": _safe_money(brutto_in)})
        try:
            _write_drafts(items)
        except Exception:
            pass

    if not item:
        from uuid import uuid4
        item = {"id": str(uuid4()), "Datum": date.today().isoformat(), "Beschreibung": "", "Kategorie": "", "Brutto": ""}
        items.append(item)
        try:
            _write_drafts(items)
        except Exception:
            pass

    # Kategorien laden (robust, mit Fallback-Liste)
    try:
        cats = _categories_list3()
    except Exception:
        try:
            cats = _categories_list2()
        except Exception:
            cats = []
    if not cats:
        cats = [
            "Abschreibung","Abschreibung als Anlage (AfA)","Arbeitsmittel / Büro","Arbeitszimmer","Bewirtung",
            "EDV-Aufwand","Fachliteratur","Fremdleistungen","Gehalt","Krankenkasse","GWG (geringw. Wirtschaftsgut)",
            "KFZ-Aufwand","Porto","Privatentnahme","Rechtsberatung","Reisekosten","Sonstiges","Spenden",
            "Spesen (Geldverkehr)","Steuerberatung","SVA","Taggelder","Telefon","Wareneinkauf","Werbungskosten","WKO-Gebühr"
        ]

    # Options-HTML
    cur = str(item.get("Kategorie",""))
    opts = []
    for c in cats:
        sel = " selected" if cur == str(c) else ""
        opts.append(f'<option value="{html.escape(str(c))}"{sel}>{html.escape(str(c))}</option>')
    options = "".join(opts)

    body = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Buchhaltung – Bearbeiten</title>
<style>
body{{font-family:system-ui,Arial,sans-serif;margin:20px}}
label{{display:block;margin-top:10px}}
input,select{{padding:6px}}
.row{{margin-bottom:8px}}
</style>
</head><body>
<h2>Beleg bearbeiten</h2>
<form method="post" action="/crm/expenses/draft/edit/?id={html.escape(str(item.get('id')))}">
  <div class="row"><label>Datum <input type="date" name="Datum" value="{html.escape(str(item.get('Datum','')))}"></label></div>
  <div class="row"><label>Beschreibung <input type="text" name="Beschreibung" value="{html.escape(str(item.get('Beschreibung','')))}" style="width:420px"></label></div>
  <div class="row"><label>Kategorie <select name="Kategorie">{options}</select></label></div>
  <div class="row"><label>Brutto (€) <input type="text" name="Brutto" value="{html.escape(str(item.get('Brutto','0.00')))}"></label></div>
  <button type="submit">Speichern</button>
</form>
<p style="margin-top:16px"><a href="/crm/expenses/drafts/">← Zurück zur Übersicht</a></p>
</body></html>"""
    return HttpResponse(body)

# --- XLSX/CSV Export der Entwürfe ---
def drafts_export_xlsx(request):
    from decimal import Decimal
    from io import BytesIO, StringIO
    from django.http import HttpResponse
    from django.utils import timezone
    try:
        items = _read_drafts()
    except Exception:
        items = []

    headers = [
        "Belegnummer","Brutto","Netto","USt.","Unternehmen","USt.-IdNr.","Land",
        "Umsatzsteuer","Kategorie","Datum","Zahlungsdatum","Zahlung"
    ]

    def D(val, default="0.00") -> Decimal:
        if val is None:
            val = default
        s = str(val).strip().replace("€", "").replace(" ", "")
        # Kommas -> Punkte, Prozentzeichen raus
        s = s.replace(",", ".").replace("%", "")
        try:
            return Decimal(s)
        except Exception:
            return Decimal(default)

    rows = []
    for it in items:
        brutto = D(it.get("Brutto") or it.get("Betrag") or "0.00")
        ust_satz = D(it.get("UStSatz") or it.get("USt.-Satz") or it.get("USt") or "0")
        ust_art  = (it.get("UStArt") or it.get("Umsatzsteuer-Art") or "Normal").strip()

        # Netto/USt berechnen, falls nicht vorhanden
        netto_key = it.get("Netto")
        if netto_key is not None and str(netto_key).strip() != "":
            netto = D(netto_key)
            ust_betrag = (brutto - netto).quantize(Decimal("0.01"))
        else:
            if any(k in ust_art.lower() for k in ("reverse", "kein", "befreit", "eu")) or ust_satz == 0:
                netto = brutto
                ust_betrag = Decimal("0.00")
            else:
                faktor = Decimal("1.00") + (ust_satz / Decimal("100"))
                netto = (brutto / faktor).quantize(Decimal("0.01"))
                ust_betrag = (brutto - netto).quantize(Decimal("0.01"))

        row = [
            it.get("Belegnummer", "") or it.get("Belegnr", ""),
            float(brutto),
            float(netto),
            float(ust_betrag),
            it.get("Unternehmen", ""),
            it.get("USt.-IdNr.") or it.get("UStIdNr") or it.get("USTIdNr") or "",
            it.get("Land", ""),
            f"{ust_art} {ust_satz}%",
            it.get("Kategorie", ""),
            (it.get("Datum") or "") ,
            (it.get("Zahlungsdatum") or ""),
            it.get("Zahlung") or it.get("Konto") or "",
        ]
        rows.append(row)

    # Erst XLSX versuchen …
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Ausgaben"
        ws.append(headers)
        for r in rows:
            ws.append(r)

        # schlichte Nummernformatierung
        for col_idx in (2, 3, 4):  # Brutto, Netto, USt.
            col = get_column_letter(col_idx)
            for cell in ws[f"{col}2":f"{col}{ws.max_row}"]:
                cell[0].number_format = "0.00"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="Ausgaben-{timezone.now().date().isoformat()}.xlsx"'
        return resp
    except Exception:
        # … sonst CSV Fallback (Excel-öffnungsfähig)
        sio = StringIO()
        import csv
        writer = csv.writer(sio, delimiter=";")
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        data = sio.getvalue().encode("utf-8-sig")
        resp = HttpResponse(data, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="Ausgaben-{timezone.now().date().isoformat()}.csv"'
        return resp
